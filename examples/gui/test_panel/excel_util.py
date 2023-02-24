"""
Author: Wei Dongcheng
Date: 2022-03-20 15:57:38
LastEditTime: 2022-03-20 16:07:39
LastEditors: Wei Dongcheng
Description:
"""
import os
import re
import subprocess
import time
import openpyxl  # pip install openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from src.myutils.file_util import FileUtil
from src.myutils.json_util import JsonUtil
from src.myutils.log_util import logger

project_path = os.path.dirname(os.path.abspath(__file__))


def cost_time(func):
    """
    func cost time decorator
    :param func: func name
    :return: wrapper
    """

    def wrapper(*args, **kwargs):
        start = time.time()
        func(*args, **kwargs)
        cost = time.time() - start
        logger.info(func.__name__ + "() cost time: " + str(round(cost, 3)) + "s")

    return wrapper


class ExcelUtil:

    @staticmethod
    def ropen_run(cmds, print_info=True):
        """
        使用subprocess.Popen执行adb命令
        :param cmds:
        :return:
        """
        if print_info:
            logger.info(" ".join(cmds))
        res = ""
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags = subprocess.CREATE_NEW_CONSOLE | subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = subprocess.SW_HIDE
        proc = subprocess.Popen(cmds,
                                stdin=subprocess.PIPE,
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                creationflags=0,
                                startupinfo=startupinfo)
        std, err = proc.communicate()
        if std is not None and std != "":
            res = std.decode(encoding="GBK")
            lines = res.replace("\r", "").split("\n")
            for line in lines:
                if line != "" and print_info is True:
                    logger.info(line.strip())
        if err is not None and err != "":
            err_info = err.decode(encoding="GBK")
            res += err_info
            lines = err_info.replace("\r", "").split("\n")
            for line in lines:
                if line != "" and print_info is True:
                    logger.error(line.strip())
        return res

    @staticmethod
    # @cost_time  # 这个会影响函数本身的返回值
    def get_col_words(excel_path, sheet_index=0, x_from="A", y_from=0):
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.worksheets[sheet_index]
        words = []
        for i in range(y_from, len(sheet[x_from])):
            value = sheet[x_from][i].value
            if value is not None:
                words.append(value)
                # LogUtil().LOGGER.info(value)
        wb.close()
        logger.info("word counter = " + str(len(words)))
        return words

    @staticmethod
    # @cost_time  # 这个会影响函数本身的返回值
    def get_words(excel_path, sheet_index=0, x_from="A", y_from=0, y_reduce=None):
        """
        获取某一列的数据
        :param excel_path:
        :param sheet_index:
        :param x_from:
        :param y_from: 起始行
        :param y_reduce: 结束行需要减去的行数，用于末尾有些行的数据不想要
        :return:
        """
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.worksheets[sheet_index]
        words = []
        if y_reduce is None:
            y_max = len(sheet[x_from])
        else:
            y_max = len(sheet[x_from]) - y_reduce
        for i in range(y_from, y_max):
            value = sheet[x_from][i].value
            if value is not None:
                words.append(str(value))
                # LogUtil().LOGGER.info(value)
        wb.close()
        logger.info("word counter = " + str(len(words)))
        return words

    @staticmethod
    def get_bug_data(sheet, bug_x="A", result_x="L", bug_id="M", bug_level="N"):
        start_time = time.time()
        counter_cases = 0
        counter_S = 0
        counter_A = 0
        counter_B = 0
        counter_C = 0
        counter_D = 0
        counter_failed_cases = 0
        if bug_x is not None and bug_x.strip() != "":
            for i in range(0, len(sheet[bug_x])):
                # print(str(column_index_from_string(bug_x)) + ",row = " + str(i))
                # value = sheet.cell(i+1, column_index_from_string(bug_x)).value
                value = sheet[bug_x][i].value
                if value is not None:
                    try:
                        int(str(value).strip())
                        # print("index value = " + str(value) + ", sheetname=" + str(sheet))
                        counter_cases = counter_cases + 1
                    except:
                        pass
        # bug level counter
        if bug_id is not None and bug_id != "" and bug_level is not None:
            for i in range(0, len(sheet[bug_id])):
                value = sheet[bug_id][i].value
                if value is not None:
                    value_level = sheet[bug_level][i].value
                    if value_level is not None and value_level.strip() != "":
                        if value_level.strip().upper() == "S":
                            logger.info(
                                "sheet name = " + str(sheet) + ", row=" + str(i) + ", level=" + str(value_level))
                            counter_S = counter_S + 1
                        elif value_level.strip().upper() == "A":
                            logger.info(
                                "sheet name = " + str(sheet) + ", row=" + str(i) + ", level=" + str(value_level))
                            counter_A = counter_A + 1
                        elif value_level.strip().upper() == "B":
                            logger.LOGGER.info(
                                "sheet name = " + str(sheet) + ", row=" + str(i) + ", level=" + str(value_level))
                            counter_B = counter_B + 1
                        elif value_level.strip().upper() == "C":
                            logger.LOGGER.info(
                                "sheet name = " + str(sheet) + ", row=" + str(i) + ", level=" + str(value_level))
                            counter_C = counter_C + 1
                        elif value_level.strip().upper() == "D":
                            logger.LOGGER.info(
                                "sheet name = " + str(sheet) + ", row=" + str(i) + ", level=" + str(value_level))
                            counter_D = counter_D + 1
        #  fail case counter
        if result_x is not None and result_x != "":
            for i in range(0, len(sheet[result_x])):
                value = sheet[result_x][i].value
                if value is not None:
                    if value.strip().upper() == "FAIL":
                        counter_failed_cases = counter_failed_cases + 1
        # wb.close()
        bug_level_counter = "S=" + str(counter_S) + ", A=" + str(counter_A) + ", B=" + str(counter_B) + ", C=" + str(
            counter_C) + ", D=" + str(counter_D)
        logger.info(bug_level_counter)
        # print("counter_cases = " + str(counter_cases))
        # print("bug_level_counter = " + str(bug_level_counter))
        # print("counter_failed_cases = " + str(counter_failed_cases))
        logger.info("cost_time = " + str(time.time() - start_time))
        return counter_cases, counter_failed_cases, counter_S, counter_A, counter_B, counter_C, counter_D

    @staticmethod
    def get_pos(sheet, bug_index="用例编号", result_x="测试结果", bug_id="对应BugID", bug_level="Bug等级"):
        logger.info("start---------------------------------------")
        bug_index_pos = None
        result_x_pos = None
        bug_id_pos = None
        bug_level_pos = None
        max_column = sheet.max_column
        max_row = sheet.max_row
        if max_column >= 50:
            max_column = 50
        if max_row >= 10:
            max_row = 10
        # value = sheet["A"][0].value  # 行，列
        # print(value)
        # value = sheet["A"][1].value  # 行，列
        # print(value)
        for j in range(1, max_column):  # j列
            for i in range(0, max_row):  # i行
                value = sheet[get_column_letter(j)][i].value  # 行，列
                if value is not None:
                    # print(str(value) + ":bug_index=" + str(bug_index) + ", result_x=" + result_x + ", bug_id=" + bug_id + ", bug_level=" + bug_level)
                    if str(value).strip() == str(bug_index).strip():
                        # print("用例编号相等")
                        bug_index_pos = get_column_letter(j)
                    elif str(value).strip() == str(result_x).strip() or str(value).strip() == "结果":
                        result_x_pos = get_column_letter(j)
                    elif str(value).strip() == str(bug_id).strip() or str(value).strip().lower() == "bugid":
                        bug_id_pos = get_column_letter(j)
                    elif str(value).strip() == str(bug_level).strip():
                        bug_level_pos = get_column_letter(j)
        logger.info("bug_index_pos = " + str(bug_index_pos))
        logger.info("result_x_pos = " + str(result_x_pos))
        logger.info("bug_id_pos = " + str(bug_id_pos))
        logger.info("bug_level_pos = " + str(bug_level_pos))
        logger.info("get_pos() end---------------------------------------")
        return bug_index_pos, result_x_pos, bug_id_pos, bug_level_pos

    @staticmethod
    def get_bug_data_by_key_words(excel_path, sheet_index=1, bug_index="用例编号", result_x="测试结果", bug_id="对应BugID",
                                  bug_level="Bug等级"):
        logger.info("path = " + excel_path)
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.worksheets[sheet_index]
        logger.info("sheet.max_row = " + str(sheet.max_row))
        logger.info("sheet.max_column = " + str(sheet.max_column))
        bug_index, result_x, bug_id, bug_level = ExcelUtil.get_pos(sheet, bug_index, result_x, bug_id, bug_level)
        counter_cases, counter_failed_cases, counter_S, counter_A, counter_B, counter_C, counter_D = \
            ExcelUtil.get_bug_data(sheet, bug_index, result_x, bug_id, bug_level)
        wb.close()
        return counter_cases, counter_failed_cases, counter_S, counter_A, counter_B, counter_C, counter_D

    @staticmethod
    def get_all_sheets_bugs(excel_path, bug_index="用例编号", result_x="测试结果", bug_id="对应BugID", bug_level="Bug等级"):
        logger.info("path = " + excel_path)
        file_cases = 0
        file_fail_case = 0
        file_s = 0
        file_a = 0
        file_b = 0
        file_c = 0
        file_d = 0
        wb = openpyxl.load_workbook(excel_path)
        for sheet_name in wb.sheetnames:
            # sheet = wb.worksheets[sheet_name]
            sheet = wb[sheet_name]
            logger.info("sheet.max_row = " + str(sheet.max_row))
            logger.info("sheet.max_column = " + str(sheet.max_column))
            bug_index_tmp, result_x_tmp, bug_id_tmp, bug_level_tmp = ExcelUtil.get_pos(sheet, bug_index, result_x,
                                                                                       bug_id, bug_level)
            counter_cases, counter_failed_cases, counter_S, counter_A, counter_B, counter_C, counter_D = \
                ExcelUtil.get_bug_data(sheet, bug_index_tmp, result_x_tmp, bug_id_tmp, bug_level_tmp)
            file_cases = file_cases + counter_cases
            file_fail_case = file_fail_case + counter_failed_cases
            file_s = file_s + counter_S
            file_a = file_a + counter_A
            file_b = file_b + counter_B
            file_c = file_c + counter_C
            file_d = file_d + counter_D
        wb.close()
        return file_cases, file_fail_case, file_s, file_a, file_b, file_c, file_d

    @staticmethod
    def get_key_word_pos(sheet, case_action="操作"):
        """
        在一定范围查找某个字符串所在的单元格
        找到返回列的字母，找不到返回None
        """
        # 操作
        cell_pos = None
        max_column = sheet.max_column
        max_row = sheet.max_row
        if max_column >= 50:
            max_column = 50
        if max_row >= 10:
            max_row = 10
        for j in range(1, max_column):  # j列
            for i in range(0, max_row):  # i行
                value = sheet[get_column_letter(j)][i].value  # 行，列, get_column_letter:获取列对应的字母
                if value is not None:
                    if str(value).strip() == str(case_action).strip():
                        cell_pos = get_column_letter(j)
        # logger.info("cell_pos = " + str(cell_pos))
        return cell_pos

    @staticmethod
    def get_all_cmds(cmd_file_path):
        start_time = time.time()
        wb = openpyxl.load_workbook(r"E:\002_智能语音\tmps\智能语音---指令.xlsx")
        sheet = wb.worksheets[0]
        value_list = []
        for i in range(1, len(sheet["J"])):
            key_value = sheet["J"][i].value
            value_list.append(str(key_value))
        wb.close()
        get_all_cmds_time = time.time() - start_time
        logger.info("cost time = " + str(round(get_all_cmds_time, 3)))
        return value_list

    @staticmethod
    @cost_time
    def get_all(dir, key_word):
        start_get_all_time = time.time()
        all_str = ""
        files_list = FileUtil.get_files_path_list(r"E:\002_智能语音\Di4.0-7.6", ".xlsx")
        for file_path in files_list:
            logger.info(file_path)
            # wb = openpyxl.load_workbook(r"D:\projects\python\gui\files\语音命令对比\Di4.0-7.6\Di4.0-Full function-7.6应用打开和关闭.xlsx")
            wb = openpyxl.load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                column_pos_letter = ExcelUtil.get_key_word_pos(sheet, "操作")  # 获取某字符串位置
                if column_pos_letter is not None:
                    for i in range(len(sheet[column_pos_letter])):
                        cell_value = sheet[column_pos_letter][i].value
                        if cell_value is not None:
                            all_str += str(cell_value)
            wb.close()
        end_get_all_time = time.time()
        cost_get_all_file_str_time = end_get_all_time - start_get_all_time
        # print(all_str)
        wb_cmd = openpyxl.load_workbook(r"E:\002_智能语音\tmps\智能语音---指令.xlsx")
        sheet_cmd = wb_cmd.worksheets[0]
        orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
        green_fill = PatternFill(start_color="AACF91", end_color="AACF91", fill_type="solid")
        stat_not_found = time.time()
        content_str = all_str.lower()
        not_found_counter = 0
        found_counter = 0
        for i in range(1, len(sheet_cmd["J"])):
            key_value = sheet_cmd["J"][i].value
            if key_value is not None:
                res = str(key_value)
                if content_str.__contains__(res.lower()):
                    sheet_cmd["J"][i].fill = green_fill
                    found_counter += 1
                else:
                    not_found_counter += 1
                    sheet_cmd["J"][i].fill = orange_fill
                    # logger.error("not found: " + res + ", counter = " + str(not_found_counter))
                    cost_not_found_time = time.time() - stat_not_found
                    # logger.info("cost_not_found_time=" + str(round(cost_not_found_time, 3)) + "--->" + res)
                    stat_not_found = time.time()
        wb_cmd.save(r"E:\002_智能语音\tmps\智能语音---指令.xlsx")
        wb_cmd.close()
        logger.info("not_found_counter = " + str(not_found_counter))
        logger.info("found_counter = " + str(found_counter))
        # print(all_str.__contains__("打开qq音乐"))

    @staticmethod
    def getPanelData(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        allPannelData = []
        for sheet_name in wb.sheetnames:
            # print(f"sheet_name = {sheet_name}")
            if sheet_name != "history":
                sheet = wb[sheet_name]
                max_column = sheet.max_column
                max_row = sheet.max_row
                dataList = []
                for i in range(1, max_row):  # i行
                    lineStr = ""
                    funcName = sheet[get_column_letter(3)][i].value
                    sigName = sheet[get_column_letter(4)][i].value
                    sigDesc = sheet[get_column_letter(5)][i].value
                    dataType = sheet[get_column_letter(7)][i].value
                    if funcName is not None and sigName is not None and sigDesc is not None and dataType is not None:
                        lineStr = f"funcName={funcName}, sigName={sigName}, sigDesc={sigDesc}, dataType={dataType}"
                        # print(lineStr)
                        dataList.append({"funcName": funcName,
                                         "sigName": sigName,
                                         "sigDesc": sigDesc,
                                         "dataType": dataType})
                allData = {"sheetname": sheet_name, "dataList": dataList}
                allPannelData.append(allData)
        return allPannelData


if __name__ == "__main__":
    logger.info("start")
    # print(ord('A'))  # char to ascii, "A" -> 65
    # ExcelUtil.get_col_words(excel_path=PROJECT_DIR + r"\files\高频话术20220317.xlsx", x_from="A", y_from=0)
    # ExcelUtil.get_col_words(excel_path=PROJECT_DIR + r"\files\chat话术.xlsx", x_from="A", y_from=1)
    # ExcelUtil.get_col_words(excel_path=PROJECT_DIR + r"\files\误唤醒测试20220329.xlsx", x_from="A", y_from=0)
    # ExcelUtil.get_col_words(excel_path=project_path + "\\优化热门中英文歌曲.xlsx", x_from="B", y_from=1)
    # ExcelUtil.get_col_words(excel_path=project_path + "\\优化热门中英文歌曲.xlsx", x_from="B", y_from=1)
    # counter_cases, bug_level_counter, counter_failed_cases = ExcelUtil.get_bug_data(r"D:\projects\python\gui\files\3.7音视频-林家声.xlsx")
    # print("用例总数 = " + str(counter_cases))
    # print("等级分布 = " + str(bug_level_counter))
    # print("失败用例 = " + str(counter_failed_cases))
    # counter_cases, counter_failed_cases, counter_S, counter_A, counter_B, counter_C, counter_D = \
    #     ExcelUtil.get_bug_data_by_key_words(r"D:\projects\python\gui\files\3.7音视频-林家声.xlsx")
    # bug_level_counter = "S=" + str(counter_S) + ", A=" + str(counter_A) + ", B=" + str(counter_B) + ", C=" + str(
    #     counter_C) + ", D=" + str(counter_D)
    # print("用例总数 = " + str(counter_cases))
    # print("等级分布 = " + str(bug_level_counter))
    # print("失败用例 = " + str(counter_failed_cases))
    # wb = openpyxl.load_workbook(r"E:\002_智能语音\003_TestReport\0322版本测试结果\Di4.0-Full function-7.6云服务安全--石云冲.xlsx")
    # print("wb.worksheets() = " + str(wb.worksheets))
    # for sheetname in wb.sheetnames:
    #     print("sheetname = " + str(sheetname))

    # LogUtil().LOGGER.info(ExcelUtil.get_all_sheets_bugs(
    #     r"/files/0322版本测试结果/Di4.0-地图导航&蓝牙电话-邱鑫堂.xlsx",
    #     bug_index="用例编号", result_x="测试结果", bug_id="对应BugID", bug_level="Bug等级"))
    # ExcelUtil.get_col_words(r"D:\projects\python\gui\files\语音命令对比\智能语音---指令.xlsx", 0, x_from="J", y_from=1)
    # ExcelUtil.get_all("", "")
    # ExcelUtil.get_semantic_task(r"E:\002_智能语音\007_自动化\GUI\semantic.xlsx")
    # name_list = ["任务1", "task2", "task3", "task4"]
    # if name_list.__contains__("task4"):
    #     print("存在哦")
    #     print(name_list.__contains__("task4"))
    #     print("index = " + str(name_list.index("task4")))
    # else:
    #     print("not exists")
    # ExcelUtil.get_excel_task(r"E:\002_智能语音\007_自动化\GUI\海外示例话术01.xlsx")
    # ExcelUtil.get_semantic_data(r"D:\projects\python\pyside6\semantic.xlsx")
    # list_tmp = ["1", "2", "3", "4", "5", "6"]
    # print(list_tmp)
    # list_tmp.clear()
    # print(list_tmp)
    res = ExcelUtil.getPanelData(r"E:\Tonsen\Download\project_lotus_interface_hud (1).xlsx")
    print(len(res))
