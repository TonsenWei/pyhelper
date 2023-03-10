# -*- coding: utf-8 -*-
"""
@Author : WeiDongcheng @tonse
@Time : 2023/3/10 15:59
@File : excel_to_xml.py
@Desc : excel转xml
pip install beautifulsoup4==4.11.1
pip install openpyxl==3.0.10
pip install lxml==4.8.0
"""
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter


def getExcelData(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    allPannelData = []
    for sheet_name in wb.sheetnames:
        # print(f"sheet_name = {sheet_name}")
        if sheet_name != "history":
            sheet = wb[sheet_name]
            max_column = sheet.max_column
            max_row = sheet.max_row
            dataList = []
            for i in range(1, max_row):  # i行
                lineStr = ""
                funcName = sheet[get_column_letter(3)][i].value
                sigName = sheet[get_column_letter(4)][i].value
                sigDesc = sheet[get_column_letter(5)][i].value
                dataType = sheet[get_column_letter(7)][i].value
                if funcName is not None and sigName is not None and sigDesc is not None and dataType is not None:
                    lineStr = f"funcName={funcName}, sigName={sigName}, sigDesc={sigDesc}, dataType={dataType}"
                    # print(lineStr)
                    dataList.append({"funcName": funcName,
                                     "sigName": sigName,
                                     "sigDesc": sigDesc,
                                     "dataType": dataType})
            allData = {"sheetname": sheet_name, "dataList": dataList}
            allPannelData.append(allData)
    return allPannelData


def data2Xml(excelData):
    xml_tree = BeautifulSoup(features="lxml-xml")

    datasourceTag = xml_tree.new_tag("datasource")
    datasourceTag.attrs = {"name": "V3_master_datasource"}
    xml_tree.append(datasourceTag)

    for i, panel in enumerate(excelData):
        sheetName = panel["sheetname"]  # [{"sheetname":"adas", "dataList":[]}]
        dataList = panel["dataList"]
        dataGroupTag = xml_tree.new_tag("datagroup")
        dataGroupTag.attrs = {"name": sheetName}
        datasourceTag.append(dataGroupTag)
        for lineCounter, lineData in enumerate(dataList):
            funcName = lineData["funcName"]
            sigName = lineData["sigName"]
            sigDesc = lineData["sigDesc"]
            dataType = lineData["dataType"]
            srcTag = xml_tree.new_tag("source")
            srcTag.attrs = {"name": sigName, "type": dataType, "value": "0"}
            dataGroupTag.append(srcTag)
    with open("file.xml", "w") as f:
        f.write(xml_tree.prettify())


if __name__ == "__main__":
    excelData = getExcelData("./hud.xlsx")
    data2Xml(excelData)
