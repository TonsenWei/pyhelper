# -*- coding: utf-8 -*-
"""
@Author : WeiDongcheng @tonse
@Time : 2023/2/6 18:24
@File : 038_openpyxl.py
@Desc : 
"""
import openpyxl
from openpyxl.utils import get_column_letter


def demo001():
    new_wb = openpyxl.Workbook()
    sheet = new_wb.worksheets[0]
    row_0 = ["用例ID", "用例名称", "测试次数", "步骤序号", "步骤类型", "参数1", "参数2", "超时时间", "测试配置", "设备选择", "步骤配置", "备注(注释)"]
    sheet.append(row_0)
    new_wb.save("测试用例.xlsx")
    new_wb.close()

def getLetter():
    print(get_column_letter(1))  # 值从1开始，返回A，B, C ……

    for i in range(1, 10):
        print(i)

def readExcel():
    wbook = openpyxl.load_workbook(r"D:\projects\python\pyside2_prjs\std_autotest\backup\export\test.xlsx")
    sheetnames = wbook.sheetnames
    taskList = []
    for sheetIndex, sheetName in enumerate(sheetnames):
        sheet = wbook[sheetName]
        for row in sheet.rows:
            if str(row[0].value) != "用例ID":
                caseIdValue = row[0].value
                caseNameValue = row[1].value
                caseTestTimesValue = row[2].value
                if caseIdValue is not None:
                    caseId = str(caseIdValue)
                    if caseNameValue is not None:
                        caseName = str(caseNameValue)
                    if caseTestTimesValue is not None:
                        caseTestTimes = str(caseTestTimesValue)
                else:  # 步骤
                    stepIdValue = row[3].value
                    stepTypeValue = row[4].value
                    stepActionValue = row[5].value
                    stepSelectedValue = row[6].value
                    stepTimeoutValue = row[7].value
                    stepOtherConfigValue = row[8].value
                    stepDeviceSelectValue = row[9].value
                    stepIsInitValue = row[10].value
                    stepMarkValue = row[11].value
        break
    wbook.close()

if __name__ == "__main__":
    readExcel()

