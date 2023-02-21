"""
Author: Wei Dongcheng
Date: 2022/8/10 13:48
LastEditTime: 2022/8/10 13:48
LastEditors: Wei Dongcheng
File: 013_logcat_analazy.py
Description:
"""
import json
import os
import re
import time

import openpyxl

from src.myutils.file_util import FileUtil


class LogcatAnalazy:
    ch_onHybridNlpResult1 = "【BYD和供应商结果都有，BYD结果胜出】"
    ch_onHybridNlpResult3 = "【等了一会，BYD和供应商结果都有，BYD结果胜出】"
    ch_onHybridNlpResult4 = "【等了一会，BYD和供应商结果都有，供应商结果胜出】"
    ch_onHybridNlpResult5 = "【等了一会，BYD没有结果，直接输出供应商结果】"
    ch_onHybridNlpResult6 = "【BYD先有结果，并直接胜出】"
    ch_onHybridNlpResult7 = "【已经有无效的BYD结果，直接输出供应商结果】"
    ch_onHybridNlpResult8 = "【供应商（讯飞）结果在白名单中，供应商（讯飞）直接胜出】"
    ch_onHybridNlpResult9 = "【BYD NLU未打开，直接使用供应商结果】"
    onHybridNlpResult1 = "HybridNlpManager: onHybridNlpResult1"
    onHybridNlpResult3 = "HybridNlpManager: onHybridNlpResult3"
    onHybridNlpResult4 = "HybridNlpManager: onHybridNlpResult4"
    onHybridNlpResult5 = "HybridNlpManager: onHybridNlpResult5"
    onHybridNlpResult6 = "HybridNlpManager: onHybridNlpResult6"
    onHybridNlpResult7 = "HybridNlpManager: onHybridNlpResult7"
    onHybridNlpResult8 = "HybridNlpManager: onHybridNlpResult8"
    onHybridNlpResult9 = "HybridNlpManager: onHybridNlpResult9"
    npl_key_word_list = [onHybridNlpResult1, onHybridNlpResult3, onHybridNlpResult4, onHybridNlpResult5,
                         onHybridNlpResult6, onHybridNlpResult7, onHybridNlpResult8, onHybridNlpResult9]
    ch_list = [ch_onHybridNlpResult1, ch_onHybridNlpResult3, ch_onHybridNlpResult4, ch_onHybridNlpResult5,
               ch_onHybridNlpResult6, ch_onHybridNlpResult7, ch_onHybridNlpResult8, ch_onHybridNlpResult9]

    @staticmethod
    def fillter_keyword(logcat_path, keyword):
        start_one_time = time.time()
        with open(logcat_path, 'r', encoding='UTF-8') as from_file:
            line = from_file.readline()
            while line:
                if line.__contains__("当前正在执行第0条"):
                    print(line)
                # if line.__contains__(keyword):
                #     try:
                #         print(line)
                #     except Exception as e:
                #         pass
                if line.__contains__("全部执行完成"):
                    print(line)
                line = from_file.readline()
        cost_time = time.time() - start_one_time
        print("耗时：" + str(round(cost_time, 3)))
        print("===============================================================================")

    @staticmethod
    def byd_vs_iflytek(logcat_path):
        start_one_time = time.time()
        byd_first_win_counter = 0
        iflytek_win_counter = 0
        byd_vendor_iflytek_win_counter = 0
        iflytek_first_win_counter = 0
        total_counter = 0
        else_counter = 0
        line_counter = 0
        file_dir, file_name = FileUtil.get_file_path_and_name(logcat_path)
        to_file_path = file_dir + "\\filter.txt"
        file_size_path = file_dir + "\\file_size.txt"
        with open(logcat_path, 'r', encoding='UTF-8') as from_file, open(to_file_path, "w+",
                                                                         encoding="UTF-8") as filter_file, open(
            file_size_path, "w+", encoding="UTF-8") as fsize_file:
            line = from_file.readline()
            while line:
                line_counter += 1
                if os.path.exists(file_size_path) and FileUtil.get_file_size(file_size_path, "m") <= 200:
                    fsize_file.write(line)
                if line.__contains__("---最终识别结果------"):
                    # if line.__contains__("---最终识别结果------") or line.__contains__("VoiceCommand") or line.__contains__(
                    #         "answer"):
                    total_counter += 1
                    # filter_file.write(line)
                if line.__contains__("【BYD先有结果，并直接胜出】"):
                    byd_first_win_counter += 1
                    print(byd_first_win_counter)
                elif line.__contains__("【已经有无效的BYD结果，直接输出供应商结果】"):
                    print(line)
                    iflytek_win_counter += 1
                elif line.__contains__("【等了一会，BYD和供应商结果都有，BYD结果胜出】"):
                    print(line)
                    byd_vendor_iflytek_win_counter += 1
                elif line.__contains__("【等了一会，BYD没有结果，直接输出供应商结果】"):
                    print(line)
                #     iflytek_first_win_counter += 1
                # elif line.__contains__("---最终识别结果------"):
                #     else_counter += 1
                #     # print("else_counter = " + str(else_counter))
                #     # print("line_counter = " + str(line_counter))
                #     try:
                #         print(str(line))
                #     except Exception as e:
                #         pass
                line = from_file.readline()
        print("===============================================================================")
        print(logcat_path)
        print("1、BYD先有结果，并直接胜出 = " + str(byd_first_win_counter))
        print("2、已经有无效的BYD结果，直接输出供应商结果 = " + str(iflytek_win_counter))
        print("3、等了一会，BYD和供应商结果都有，BYD结果胜出 = " + str(byd_vendor_iflytek_win_counter))
        print("4、等了一会，BYD没有结果，直接输出供应商结果 = " + str(iflytek_first_win_counter))
        print("总数：" + str(total_counter))
        print("[1]BYD先有结果，并直接胜出占比：" + str(byd_first_win_counter / total_counter))
        print("[2]已经有无效的BYD结果，直接输出供应商结果占比 = " + str(iflytek_win_counter / total_counter))
        print("[3]等了一会，BYD和供应商结果都有，BYD结果胜出占比 = " + str(byd_vendor_iflytek_win_counter / total_counter))
        print("[4]等了一会，BYD没有结果，直接输出供应商结果占比 = " + str(iflytek_first_win_counter / total_counter))
        cost_time = time.time() - start_one_time
        print("耗时：" + str(round(cost_time, 3)) + ", 文件行数：" + str(line_counter))
        print("===============================================================================")
        print("")
        return byd_first_win_counter, iflytek_win_counter, byd_vendor_iflytek_win_counter, iflytek_first_win_counter, total_counter

    @staticmethod
    def byd_vs_iflytek_multi(path_list):
        """
        多文件解析
        """
        start_time = time.time()
        all_byd_first_win_counter = 0
        all_iflytek_win_counter = 0
        all_byd_vendor_iflytek_win_counter = 0
        all_iflytek_first_win_counter = 0
        all_total_counter = 0
        for log_path in path_list:
            byd_win_first, ifl_win, byd_vendor, ifl_win_first, total = LogcatAnalazy.byd_vs_iflytek(log_path)
            all_byd_first_win_counter += byd_win_first
            all_iflytek_win_counter += ifl_win
            all_byd_vendor_iflytek_win_counter += byd_vendor
            all_iflytek_first_win_counter += ifl_win_first
            all_total_counter += total
        print("total=========================================================================")
        print("1、BYD先有结果，并直接胜出 = " + str(all_byd_first_win_counter))
        print("2、已经有无效的BYD结果，直接输出供应商结果 = " + str(all_iflytek_win_counter))
        print("3、等了一会，BYD和供应商结果都有，BYD结果胜出 = " + str(all_byd_vendor_iflytek_win_counter))
        print("4、等了一会，BYD没有结果，直接输出供应商结果 = " + str(all_iflytek_first_win_counter))
        print("总数：" + str(all_total_counter))
        print("[1]BYD先有结果，并直接胜出占比：" + str(all_byd_first_win_counter / all_total_counter))
        print("[2]已经有无效的BYD结果，直接输出供应商结果占比 = " + str(all_iflytek_win_counter / all_total_counter))
        print("[3]等了一会，BYD和供应商结果都有，BYD结果胜出占比 = " + str(all_byd_vendor_iflytek_win_counter / all_total_counter))
        print("[4]等了一会，BYD没有结果，直接输出供应商结果占比 = " + str(all_iflytek_first_win_counter / all_total_counter))
        cost_time = time.time() - start_time
        print("耗时：" + str(round(cost_time, 3)))

    @staticmethod
    def byd_vs_iflytek_001(logcat_path):
        start_one_time = time.time()
        npl001_counter = 0
        npl003_counter = 0
        npl004_counter = 0
        npl005_counter = 0
        npl006_counter = 0
        npl007_counter = 0
        npl008_counter = 0
        npl009_counter = 0
        else_counter = 0
        total_counter = 0
        line_counter = 0
        file_dir, file_name = FileUtil.get_file_path_and_name(logcat_path)
        tmp_file_path = file_dir + "\\tmp.txt"
        with open(logcat_path, 'r', encoding='UTF-8') as from_file, open(tmp_file_path, "a+",
                                                                         encoding="UTF-8") as tmp_file:
            line = from_file.readline()
            while line:
                line_counter += 1
                if line.__contains__("---最终识别结果------"):
                    total_counter += 1
                    # filter_file.write(line)
                if line.__contains__(LogcatAnalazy.onHybridNlpResult1):
                    npl001_counter += 1
                elif line.__contains__(LogcatAnalazy.onHybridNlpResult3):
                    # print(line)
                    npl003_counter += 1
                elif line.__contains__(LogcatAnalazy.onHybridNlpResult4):  # onHybridNlpResult4
                    npl004_counter += 1
                elif line.__contains__(LogcatAnalazy.onHybridNlpResult5):  # onHybridNlpResult5
                    npl005_counter += 1
                    # print(line)
                elif line.__contains__(LogcatAnalazy.onHybridNlpResult6):
                    npl006_counter += 1
                    # print(byd_first_win_counter)
                elif line.__contains__(LogcatAnalazy.onHybridNlpResult7):  # onHybridNlpResult7
                    # print(line)
                    npl007_counter += 1
                elif line.__contains__(LogcatAnalazy.onHybridNlpResult8):
                    npl008_counter += 1
                elif line.__contains__(LogcatAnalazy.onHybridNlpResult9):
                    npl009_counter += 1
                    # print(line)
                elif line.__contains__("HybridNlpManager: onHybridNlpResult"):
                    else_counter += 1
                    try:
                        tmp_file.write(line)
                        print("=================" + line)
                    except Exception as e:
                        print("error: " + str(e))
                line = from_file.readline()
        print("===============================================================================")
        print(logcat_path)
        npl_counter_list = [npl001_counter, npl003_counter, npl004_counter, npl005_counter,
                            npl006_counter, npl007_counter, npl008_counter, npl009_counter]
        key_word_index = -1
        for per_nlp in LogcatAnalazy.ch_list:
            key_word_index += 1
            print(str(key_word_index) + "、" + per_nlp + "=" + str(
                npl_counter_list[key_word_index]))
        print("总数：" + str(total_counter))
        key_word_index = -1
        for per_nlp in LogcatAnalazy.ch_list:
            key_word_index += 1
            print("[" + str(key_word_index) + "]" + per_nlp + " 占比=" + str(
                npl_counter_list[key_word_index] / total_counter * 100) + "%")
        cost_time = time.time() - start_one_time
        print("耗时：" + str(round(cost_time, 3)) + ", 文件行数：" + str(line_counter))
        print("===============================================================================")
        return npl_counter_list, total_counter, else_counter

    @staticmethod
    def byd_vs_iflytek_multi_001(path_list):
        """
        多文件解析
        """
        start_time = time.time()
        mul_npl001_counter = 0
        mul_npl003_counter = 0
        mul_npl004_counter = 0
        mul_npl005_counter = 0
        mul_npl006_counter = 0
        mul_npl007_counter = 0
        mul_npl008_counter = 0
        mul_npl009_counter = 0
        mul_else_counter = 0
        all_counter = 0
        for log_path in path_list:
            npl_counter_list, total_counter, else_counter = LogcatAnalazy.byd_vs_iflytek_001(
                log_path)
            mul_npl001_counter += npl_counter_list[0]
            mul_npl003_counter += npl_counter_list[1]
            mul_npl004_counter += npl_counter_list[2]
            mul_npl005_counter += npl_counter_list[3]
            mul_npl006_counter += npl_counter_list[4]
            mul_npl007_counter += npl_counter_list[5]
            mul_npl008_counter += npl_counter_list[6]
            mul_npl009_counter += npl_counter_list[7]
            all_counter += total_counter
            mul_else_counter += else_counter
        mul_npl_counter_list = [mul_npl001_counter, mul_npl003_counter, mul_npl004_counter, mul_npl005_counter,
                                mul_npl006_counter, mul_npl007_counter, mul_npl008_counter, mul_npl009_counter]
        key_word_index = -1
        print("[.............................多文件解析结果.............................]")
        for per_nlp in LogcatAnalazy.ch_list:
            key_word_index += 1
            print(str(key_word_index) + "、" + per_nlp + "=" + str(
                mul_npl_counter_list[key_word_index]))
        key_word_index += 1
        print(str(key_word_index) + "、else = " + str(mul_else_counter))
        real_all_counter = 0
        for mul_num in mul_npl_counter_list:
            real_all_counter += mul_num
        real_all_counter += mul_else_counter
        # print("总数：" + str(all_counter) + ", 解析成功总数：" + str(real_all_counter))
        print("解析成功总数：" + str(real_all_counter))
        key_word_index = -1
        for per_nlp in LogcatAnalazy.ch_list:
            key_word_index += 1
            percent = round(mul_npl_counter_list[key_word_index] / real_all_counter * 100, 3)
            print("[" + str(key_word_index) + "]" + per_nlp + " 占比=" + str(percent) + "%")
        percent = round(mul_else_counter / real_all_counter * 100, 3)
        key_word_index += 1
        print("[" + str(key_word_index) + "] else 占比= " + str(percent) + "%")
        cost_time = time.time() - start_time
        print("总耗时：" + str(round(cost_time, 3)))

    @staticmethod
    def byd_vs_iflytek_get_data(logcat_path):
        """
        单个logcat.txt分析并输出数据到txt和excel(xlsx),用于跑了单个或多个任务但是只输出日志到单个logcat的情况
        :param logcat_path: logcat日志路径
        :return:
        """
        record_list = []
        file_dir, file_name = FileUtil.get_file_path_and_name(logcat_path)
        to_file_path = file_dir + "\\data.txt"
        excel_path = file_dir + "\\data.xlsx"
        new_wb = openpyxl.Workbook()
        default_sheet = new_wb.worksheets[0]
        row_index = 0
        row_index += 1
        default_sheet["A" + str(row_index)] = "结果"
        default_sheet["B" + str(row_index)] = "指令"
        default_sheet["C" + str(row_index)] = "回答"
        default_sheet["D" + str(row_index)] = "cmdId"
        default_sheet["E" + str(row_index)] = "commandId"
        # 单元格宽度
        default_sheet.column_dimensions["A"].width = 32
        default_sheet.column_dimensions["B"].width = 32
        default_sheet.column_dimensions["C"].width = 32
        default_sheet.column_dimensions["D"].width = 16
        default_sheet.column_dimensions["E"].width = 16
        with open(logcat_path, 'r', encoding='UTF-8') as from_file, open(to_file_path, "w+",
                                                                         encoding='UTF-8') as to_file:
            line = from_file.readline()
            record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "", "commandId": ""}
            while line:
                if line.__contains__("---最终识别结果------"):  # 数据开始，初始化
                    record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                   "commandId": ""}
                    target = line.split("【")[1].replace("】", "").replace("\n", "")
                    record_data["res_type"] = target
                elif line.__contains__("VoiceCommand="):
                    target = line.split("VoiceCommand=")[1].replace("\n", "")
                    record_data["VoiceCommand"] = target
                elif line.__contains__("BYD_AUTO_TESTING_: audio_recognize:"):
                    answers_slot = ""
                    cmdId_list = re.findall("cmdId=[0-9]*", line)
                    if len(cmdId_list) == 1:
                        cmdId_str = cmdId_list[0]
                        cmdId_str_value = cmdId_str.split("=")[1]
                        record_data["cmdId"] = cmdId_str_value
                    first_slot_split = line.split("slot=")
                    if len(first_slot_split) == 2:
                        answers_slot += first_slot_split[1]
                    try:
                        record_data["audio_recognize"] = json.loads(answers_slot)["answer"]
                    except Exception as e:
                        # {"action":"","answer":"屏幕亮度已调到0%","answerNum":
                        ans_list = answers_slot.split("\"answer\":\"")
                        if len(ans_list) == 2:
                            record_data["audio_recognize"] = ans_list[1].split("\",")[0]
                        else:
                            record_data["audio_recognize"] = answers_slot
                        print("commandId: " + str(e))
                    try:
                        record_data["commandId"] = json.loads(answers_slot)["commandId"]
                    except Exception as e:
                        # "color":"","commandId":109272,"dialogScene":0,
                        ans_list = answers_slot.split("\"commandId\":\"")
                        if len(ans_list) == 2:
                            record_data["commandId"] = ans_list[1].split(",")[0]
                        else:
                            record_data["commandId"] = ""
                        print("commandId: " + str(e))
                    if record_data["res_type"] != "" and record_data["VoiceCommand"] != "" and record_data[
                        "audio_recognize"] != "":
                        to_file.write(str(record_data) + "\n")
                        record_list.append(record_data)
                        row_index += 1
                        default_sheet["A" + str(row_index)] = record_data["res_type"]
                        default_sheet["B" + str(row_index)] = record_data["VoiceCommand"]
                        default_sheet["C" + str(row_index)] = record_data["audio_recognize"]
                        default_sheet["D" + str(row_index)] = record_data["cmdId"]
                        default_sheet["E" + str(row_index)] = record_data["commandId"]

                        # 一条数据结束，重新初始化数据
                        record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                       "commandId": ""}
                line = from_file.readline()
        print("record_list size: " + str(len(record_list)))
        new_wb.save(excel_path)
        new_wb.close()

    @staticmethod
    def byd_vs_iflytek_get_data_multi(path_list, data_dir):
        """
        多个logcat.txt分析并输出数据到txt和excel(xlsx),用于跑了多个任务，每个任务生成一个logcat.txt的情况
        多个logcat分析出来的数据最后会放到一起输出
        :param path_list: 文件路径列表
        :param data_dir: 数据保存目录
        :return:
        """
        record_list = []
        to_file_path = data_dir + "\\data.txt"
        excel_path = data_dir + "\\data.xlsx"

        new_wb = openpyxl.Workbook()
        default_sheet = new_wb.worksheets[0]
        row_index = 0
        row_index += 1
        default_sheet["A" + str(row_index)] = "结果"
        default_sheet["B" + str(row_index)] = "指令"
        default_sheet["C" + str(row_index)] = "回答"
        default_sheet["D" + str(row_index)] = "cmdId"
        default_sheet["E" + str(row_index)] = "commandId"
        # 单元格宽度
        default_sheet.column_dimensions["A"].width = 32
        default_sheet.column_dimensions["B"].width = 32
        default_sheet.column_dimensions["C"].width = 32
        default_sheet.column_dimensions["D"].width = 16
        default_sheet.column_dimensions["E"].width = 16

        for per_path in path_list:
            print("正在解析：" + per_path)
            with open(to_file_path, "w+", encoding='UTF-8') as to_file:
                with open(per_path, 'r', encoding='UTF-8') as from_file:
                    line = from_file.readline()
                    record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                   "commandId": ""}
                    while line:
                        if line.__contains__("---最终识别结果------"):  # 数据开始，初始化
                            record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                           "commandId": ""}
                            target = line.split("【")[1].replace("】", "").replace("\n", "")
                            record_data["res_type"] = target
                        elif line.__contains__("VoiceCommand="):
                            target = line.split("VoiceCommand=")[1].replace("\n", "")
                            record_data["VoiceCommand"] = target
                        elif line.__contains__("BYD_AUTO_TESTING_: audio_recognize:"):
                            answers_slot = ""
                            cmdId_list = re.findall("cmdId=[0-9]*", line)
                            if len(cmdId_list) == 1:
                                cmdId_str = cmdId_list[0]
                                cmdId_str_value = cmdId_str.split("=")[1]
                                record_data["cmdId"] = cmdId_str_value
                            first_slot_split = line.split("slot=")
                            if len(first_slot_split) == 2:
                                answers_slot += first_slot_split[1]
                            try:
                                record_data["audio_recognize"] = json.loads(answers_slot)["answer"]
                            except Exception as e:
                                # {"action":"","answer":"屏幕亮度已调到0%","answerNum":
                                ans_list = answers_slot.split("\"answer\":\"")
                                if len(ans_list) == 2:
                                    record_data["audio_recognize"] = ans_list[1].split("\",")[0]
                                else:
                                    record_data["audio_recognize"] = answers_slot
                                print("audio_recognize: " + str(e))
                            try:
                                record_data["commandId"] = json.loads(answers_slot)["commandId"]
                            except Exception as e:
                                # "color":"","commandId":109272,"dialogScene":0,
                                ans_list = answers_slot.split("\"commandId\":\"")
                                if len(ans_list) == 2:
                                    record_data["commandId"] = ans_list[1].split(",")[0]
                                else:
                                    record_data["commandId"] = ""
                                print("commandId: " + str(e))
                            if record_data["res_type"] != "" and record_data["VoiceCommand"] != "" and record_data[
                                "audio_recognize"] != "":
                                to_file.write(str(record_data) + "\n")
                                record_list.append(record_data)
                                row_index += 1
                                default_sheet["A" + str(row_index)] = record_data["res_type"]
                                default_sheet["B" + str(row_index)] = record_data["VoiceCommand"]
                                default_sheet["C" + str(row_index)] = record_data["audio_recognize"]
                                default_sheet["D" + str(row_index)] = record_data["cmdId"]
                                default_sheet["E" + str(row_index)] = record_data["commandId"]

                                # 一条数据结束，重新初始化数据
                                record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                               "commandId": ""}
                        line = from_file.readline()
        print("record_list size: " + str(len(record_list)))
        new_wb.save(excel_path)
        new_wb.close()

    @staticmethod
    def byd_vs_iflytek_get_data_001(logcat_path):
        """
        单个logcat.txt分析并输出数据到txt和excel(xlsx),用于跑了单个或多个任务但是只输出日志到单个logcat的情况
        :param logcat_path: logcat日志路径
        :return:
        """
        record_list = []
        file_dir, file_name = FileUtil.get_file_path_and_name(logcat_path)
        to_file_path = file_dir + "\\data.txt"
        excel_path = file_dir + "\\data.xlsx"
        new_wb = openpyxl.Workbook()
        default_sheet = new_wb.worksheets[0]
        row_index = 0
        row_index += 1
        default_sheet["A" + str(row_index)] = "结果"
        default_sheet["B" + str(row_index)] = "指令"
        default_sheet["C" + str(row_index)] = "回答"
        default_sheet["D" + str(row_index)] = "cmdId"
        default_sheet["E" + str(row_index)] = "commandId"
        # 单元格宽度
        default_sheet.column_dimensions["A"].width = 32
        default_sheet.column_dimensions["B"].width = 32
        default_sheet.column_dimensions["C"].width = 32
        default_sheet.column_dimensions["D"].width = 16
        default_sheet.column_dimensions["E"].width = 16
        with open(logcat_path, 'r', encoding='UTF-8') as from_file, open(to_file_path, "w+",
                                                                         encoding='UTF-8') as to_file:
            line = from_file.readline()
            record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "", "commandId": ""}
            while line:
                if line.__contains__("---最终识别结果------"):  # 数据开始，初始化
                    record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                   "commandId": ""}
                    target = line.split("【")[1].replace("】", "").replace("\n", "")
                    record_data["res_type"] = target
                elif line.__contains__("VoiceCommand="):
                    target = line.split("VoiceCommand=")[1].replace("\n", "")
                    record_data["VoiceCommand"] = target
                elif line.__contains__("BYD_AUTO_TESTING_: audio_recognize:"):
                    answers_slot = ""
                    cmdId_list = re.findall("cmdId=[0-9]*", line)
                    if len(cmdId_list) == 1:
                        cmdId_str = cmdId_list[0]
                        cmdId_str_value = cmdId_str.split("=")[1]
                        record_data["cmdId"] = cmdId_str_value
                    first_slot_split = line.split("slot=")
                    if len(first_slot_split) == 2:
                        answers_slot += first_slot_split[1]
                    try:
                        record_data["audio_recognize"] = json.loads(answers_slot)["answer"]
                    except Exception as e:
                        # {"action":"","answer":"屏幕亮度已调到0%","answerNum":
                        ans_list = answers_slot.split("\"answer\":\"")
                        if len(ans_list) == 2:
                            record_data["audio_recognize"] = ans_list[1].split("\",")[0]
                        else:
                            record_data["audio_recognize"] = answers_slot
                        print("commandId: " + str(e))
                    try:
                        record_data["commandId"] = json.loads(answers_slot)["commandId"]
                    except Exception as e:
                        # "color":"","commandId":109272,"dialogScene":0,
                        ans_list = answers_slot.split("\"commandId\":\"")
                        if len(ans_list) == 2:
                            record_data["commandId"] = ans_list[1].split(",")[0]
                        else:
                            record_data["commandId"] = ""
                        print("commandId: " + str(e))
                    if record_data["res_type"] != "" and record_data["VoiceCommand"] != "" and record_data[
                        "audio_recognize"] != "":
                        to_file.write(str(record_data) + "\n")
                        record_list.append(record_data)
                        row_index += 1
                        default_sheet["A" + str(row_index)] = record_data["res_type"]
                        default_sheet["B" + str(row_index)] = record_data["VoiceCommand"]
                        default_sheet["C" + str(row_index)] = record_data["audio_recognize"]
                        default_sheet["D" + str(row_index)] = record_data["cmdId"]
                        default_sheet["E" + str(row_index)] = record_data["commandId"]

                        # 一条数据结束，重新初始化数据
                        record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                       "commandId": ""}
                line = from_file.readline()
        print("record_list size: " + str(len(record_list)))
        new_wb.save(excel_path)
        new_wb.close()

    @staticmethod
    def byd_vs_iflytek_get_data_multi_001(path_list, data_dir):
        """
        多个logcat.txt分析并输出数据到txt和excel(xlsx),用于跑了多个任务，每个任务生成一个logcat.txt的情况
        多个logcat分析出来的数据最后会放到一起输出
        :param path_list: 文件路径列表
        :param data_dir: 数据保存目录
        :return:
        """
        record_list = []
        to_file_path = data_dir + "\\data.txt"
        excel_path = data_dir + "\\data.xlsx"

        new_wb = openpyxl.Workbook()
        default_sheet = new_wb.worksheets[0]
        row_index = 0
        row_index += 1
        default_sheet["A" + str(row_index)] = "结果"
        default_sheet["B" + str(row_index)] = "指令"
        default_sheet["C" + str(row_index)] = "回答"
        default_sheet["D" + str(row_index)] = "cmdId"
        default_sheet["E" + str(row_index)] = "commandId"
        # 单元格宽度
        default_sheet.column_dimensions["A"].width = 32
        default_sheet.column_dimensions["B"].width = 32
        default_sheet.column_dimensions["C"].width = 32
        default_sheet.column_dimensions["D"].width = 16
        default_sheet.column_dimensions["E"].width = 16

        for per_path in path_list:
            print("正在解析：" + per_path)
            with open(to_file_path, "w+", encoding='UTF-8') as to_file:
                with open(per_path, 'r', encoding='UTF-8') as from_file:
                    line = from_file.readline()
                    record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                   "commandId": ""}
                    while line:
                        if line.__contains__("HybridNlpManager: onHybridNlpResult"):  # 数据开始，初始化
                            record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                           "commandId": ""}
                            target = line.split("【")[1].replace("】", "").replace("\n", "")
                            record_data["res_type"] = target
                        elif line.__contains__("VoiceCommand="):
                            target = line.split("VoiceCommand=")[1].replace("\n", "")
                            record_data["VoiceCommand"] = target
                        elif line.__contains__("BYD_AUTO_TESTING_: audio_recognize:"):
                            answers_slot = ""
                            # try:
                            #     print(line)
                            # except Exception as e:
                            #     pass
                            cmdId_list = re.findall("cmdId=[0-9]*", line)
                            if len(cmdId_list) == 1:
                                cmdId_str = cmdId_list[0]
                                cmdId_str_value = cmdId_str.split("=")[1]
                                record_data["cmdId"] = cmdId_str_value
                            first_slot_split = line.split("slot=")
                            if len(first_slot_split) == 2:
                                answers_slot += first_slot_split[1]
                            try:
                                record_data["audio_recognize"] = json.loads(answers_slot)["answer"]
                            except Exception as e:
                                # {"action":"","answer":"屏幕亮度已调到0%","answerNum":
                                ans_list = answers_slot.split("\"answer\":\"")
                                if len(ans_list) == 2:
                                    record_data["audio_recognize"] = ans_list[1].split("\",")[0]
                                else:
                                    record_data["audio_recognize"] = answers_slot
                                print("audio_recognize except: " + str(e))
                            try:
                                record_data["commandId"] = json.loads(answers_slot)["commandId"]
                            except Exception as e:
                                # "color":"","commandId":109272,"dialogScene":0,
                                ans_list = answers_slot.split("\"commandId\":\"")
                                if len(ans_list) == 2:
                                    record_data["commandId"] = ans_list[1].split(",")[0]
                                else:
                                    record_data["commandId"] = ""
                                print("commandId except: " + str(e))
                            if record_data["res_type"] != "" and record_data["VoiceCommand"] != "" and record_data["audio_recognize"] != "":
                                to_file.write(str(record_data) + "\n")
                                record_list.append(record_data)
                                row_index += 1
                                default_sheet["A" + str(row_index)] = record_data["res_type"]
                                default_sheet["B" + str(row_index)] = record_data["VoiceCommand"]
                                default_sheet["C" + str(row_index)] = record_data["audio_recognize"]
                                default_sheet["D" + str(row_index)] = record_data["cmdId"]
                                default_sheet["E" + str(row_index)] = record_data["commandId"]

                                # 一条数据结束，重新初始化数据
                                record_data = {"res_type": "", "VoiceCommand": "", "audio_recognize": "", "cmdId": "",
                                               "commandId": ""}
                        line = from_file.readline()
        print("record_list size: " + str(len(record_list)))
        new_wb.save(excel_path)
        new_wb.close()


if __name__ == "__main__":
    # file_path0 = r"D:\software\YY\20220810_01语义测试\20220628008\logcat.txt"     # 待分析logcat路径
    # file_path = r"D:\software\YY\20220810_01语义测试\20220628008\logcat.txt"     # 待分析logcat路径
    # file_path002 = r"D:\software\YY\20220810_01语义测试\20220810002\logcat.txt"  # 待分析logcat路径
    # file_path003 = r"D:\software\YY\20220810_01语义测试\20220810003\logcat.txt"  # 待分析logcat路径
    # path_list = [file_path, file_path002, file_path003]

    lg_007_0 = r"E:\logs\2022-08-25_语义自动化\2022-08-25_12-25-02_724_logcat_007.txt"
    lg_007 = r"E:\logs\2022-08-25_语义自动化\2022-08-25_17-49-41_607_logcat_007.txt"
    lg_008 = r"E:\logs\2022-08-25_语义自动化\2022-08-25_16-28-39_473_logcat_008.txt"
    lg_009 = r"E:\logs\2022-08-25_语义自动化\2022-08-25_15-13-57_127_logcat_009.txt"
    lg_010 = r"E:\logs\2022-08-25_语义自动化\2022-08-25_14-09-57_473_logcat_010.txt"
    lg_011 = r"E:\logs\2022-08-25_语义自动化\2022-08-25_11-51-17_065_logcat_011.txt"
    # path_list = [lg_007_0]
    path_list = [lg_007, lg_008, lg_009, lg_010, lg_011]
    # LogcatAnalazy.byd_vs_iflytek(lg_007_0)
    # LogcatAnalazy.byd_vs_iflytek_001(lg_007_0)

    # 跑批结果解析
    # LogcatAnalazy.byd_vs_iflytek_multi_001(path_list)
    # LogcatAnalazy.byd_vs_iflytek_multi(path_list)

    # LogcatAnalazy.fillter_keyword(r"E:\logs\2022-08-25_语义自动化\2022-08-25_14-09-57_473_logcat_010.txt", "automatedTasksState")
    # path_list = [file_path]
    # LogcatAnalazy.byd_vs_iflytek_multi(path_list)
    # LogcatAnalazy.byd_vs_iflytek_get_data(file_path)

    # 从跑批中提取指令和commandId信息
    # LogcatAnalazy.byd_vs_iflytek_get_data_multi(path_list, r"D:\software\YY\20220810_01语义测试\output\20220811_01")
    LogcatAnalazy.byd_vs_iflytek_get_data_multi_001(path_list, r"E:\logs\2022-08-25_语义自动化\提取语义")
    print("end! <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")
