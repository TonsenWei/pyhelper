"""
Author: Wei Dongcheng
Date: 2022-03-20 15:57:38
LastEditTime: 2022-03-20 16:07:39
LastEditors: Wei Dongcheng
Description:
"""
import os
import openpyxl  # pip install openpyxl
from openpyxl.utils import get_column_letter


project_path = os.path.dirname(os.path.abspath(__file__))


class ExcelUtil:

    @staticmethod
    def getPanelData(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        allPannelData = []
        for sheet_name in wb.sheetnames:
            # print(f"sheet_name = {sheet_name}")
            if sheet_name != "history":
                sheet = wb[sheet_name]
                max_column = sheet.max_column
                max_row = sheet.max_row
                dataList = []
                for i in range(1, max_row):  # i行
                    lineStr = ""
                    funcName = sheet[get_column_letter(3)][i].value
                    sigName = sheet[get_column_letter(4)][i].value
                    sigDesc = sheet[get_column_letter(5)][i].value
                    dataType = sheet[get_column_letter(7)][i].value
                    if funcName is not None and sigName is not None and sigDesc is not None and dataType is not None:
                        lineStr = f"funcName={funcName}, sigName={sigName}, sigDesc={sigDesc}, dataType={dataType}"
                        # print(lineStr)
                        dataList.append({"funcName": funcName,
                                         "sigName": sigName,
                                         "sigDesc": sigDesc,
                                         "dataType": dataType})
                allData = {"sheetname": sheet_name, "dataList": dataList}
                allPannelData.append(allData)
        return allPannelData

